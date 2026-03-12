"""Helpers for turning uploaded data files into prompt context."""

import csv
import json
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook

from .config import settings

SUPPORTED_UPLOAD_SUFFIXES = {".csv", ".json", ".md", ".txt", ".xlsx"}


class DataSourceError(ValueError):
    """Raised when an uploaded file cannot be summarized safely."""


def summarize_file_bytes(
    filename: str,
    content: bytes,
    media_type: str | None = None,
) -> dict:
    """Summarize one uploaded file into a compact prompt context block."""
    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_UPLOAD_SUFFIXES:
        raise DataSourceError(
            f"Unsupported file type '{suffix or 'unknown'}'. Use CSV, JSON, XLSX, TXT, or MD."
        )
    if len(content) > settings.max_upload_bytes:
        raise DataSourceError(
            f"{filename} is too large. Limit uploads to "
            f"{settings.max_upload_bytes // 1_000_000} MB."
        )

    if suffix == ".csv":
        return _summarize_csv(filename, content, media_type)
    if suffix == ".json":
        return _summarize_json(filename, content, media_type)
    if suffix == ".xlsx":
        return _summarize_xlsx(filename, content, media_type)
    return _summarize_text(filename, content, media_type)


def summarize_path(path: Path) -> dict:
    """Summarize a local file path for CLI usage."""
    return summarize_file_bytes(path.name, path.read_bytes())


def _summarize_csv(filename: str, content: bytes, media_type: str | None) -> dict:
    text = content.decode("utf-8-sig", errors="replace")
    rows = list(csv.reader(StringIO(text)))
    if not rows:
        raise DataSourceError(f"{filename} is empty.")

    header = [str(cell).strip() for cell in rows[0] if str(cell).strip()]
    sample_rows = rows[1:4]
    summary = [
        f"File: {filename}",
        "Type: CSV table",
        f"Rows: {max(len(rows) - 1, 0)}",
        f"Columns: {', '.join(header) if header else 'No header row detected'}",
    ]

    if sample_rows:
        summary.append("Sample rows:")
        summary.extend(
            f"- {', '.join(str(cell) for cell in row[:8])}" for row in sample_rows if any(row)
        )

    return {
        "filename": filename,
        "file_type": "csv",
        "media_type": media_type,
        "summary": "\n".join(summary),
        "suggested_diagram_types": _suggest_types_from_header(header),
    }


def _summarize_json(filename: str, content: bytes, media_type: str | None) -> dict:
    data = json.loads(content.decode("utf-8"))
    summary = [f"File: {filename}", "Type: JSON document"]
    suggestions = ["flowchart"]

    if isinstance(data, list):
        summary.append(f"Top-level structure: array with {len(data)} item(s)")
        if data and isinstance(data[0], dict):
            keys = sorted({key for item in data[:10] if isinstance(item, dict) for key in item})
            summary.append(f"Observed keys: {', '.join(keys) if keys else 'None'}")
            summary.append("Sample records:")
            for item in data[:3]:
                summary.append(f"- {json.dumps(item, ensure_ascii=True)[:220]}")
            suggestions = _suggest_types_from_header(keys)
        else:
            summary.append(f"Sample values: {json.dumps(data[:3], ensure_ascii=True)[:220]}")
    elif isinstance(data, dict):
        keys = list(data.keys())[:12]
        summary.append(
            f"Top-level structure: object with keys {', '.join(keys) if keys else 'none'}"
        )
        sample = {key: data[key] for key in keys[:3]}
        summary.append(f"Sample values: {json.dumps(sample, ensure_ascii=True)[:220]}")
        suggestions = _suggest_types_from_header(keys)
    else:
        summary.append(f"Value: {json.dumps(data, ensure_ascii=True)[:220]}")

    return {
        "filename": filename,
        "file_type": "json",
        "media_type": media_type,
        "summary": "\n".join(summary),
        "suggested_diagram_types": suggestions,
    }


def _summarize_xlsx(filename: str, content: bytes, media_type: str | None) -> dict:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    summary = [f"File: {filename}", f"Type: XLSX workbook with {len(workbook.sheetnames)} sheet(s)"]
    suggestions: set[str] = {"flowchart"}

    for sheet_name in workbook.sheetnames[:3]:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(min_row=1, max_row=4, values_only=True))
        header = [
            str(cell).strip()
            for cell in (rows[0] if rows else [])
            if cell is not None and str(cell).strip()
        ]
        suggestions.update(_suggest_types_from_header(header))
        summary.append(
            f"Sheet '{sheet_name}': {max(sheet.max_row - 1, 0)} data row(s), "
            f"{sheet.max_column} column(s)"
        )
        if header:
            summary.append(f"- Columns: {', '.join(header)}")
        for sample in rows[1:4]:
            values = [str(cell) for cell in sample if cell is not None][:8]
            if values:
                summary.append(f"- Sample: {', '.join(values)}")

    workbook.close()
    return {
        "filename": filename,
        "file_type": "xlsx",
        "media_type": media_type,
        "summary": "\n".join(summary),
        "suggested_diagram_types": sorted(suggestions),
    }


def _summarize_text(filename: str, content: bytes, media_type: str | None) -> dict:
    text = content.decode("utf-8", errors="replace").strip()
    excerpt = "\n".join(line for line in text.splitlines()[:8] if line.strip())
    if not excerpt:
        raise DataSourceError(f"{filename} is empty.")

    return {
        "filename": filename,
        "file_type": Path(filename).suffix.lower().lstrip("."),
        "media_type": media_type,
        "summary": f"File: {filename}\nType: Text document\nExcerpt:\n{excerpt[:1200]}",
        "suggested_diagram_types": ["flowchart"],
    }


def combine_summaries(items: list[dict]) -> str:
    """Merge multiple summaries into one prompt block."""
    return "\n\n".join(item["summary"] for item in items)


def _suggest_types_from_header(header: list[str]) -> list[str]:
    joined = " ".join(column.lower() for column in header)
    suggestions = {"flowchart"}

    if any(keyword in joined for keyword in ("date", "month", "year", "revenue", "count", "sales")):
        suggestions.add("chart")
    if any(keyword in joined for keyword in ("source", "target", "from", "to", "depends")):
        suggestions.add("flowchart")
    if any(keyword in joined for keyword in ("service", "system", "component", "owner")):
        suggestions.add("architecture")
    if any(keyword in joined for keyword in ("subnet", "switch", "router", "firewall")):
        suggestions.add("network")

    return sorted(suggestions)
